from collections import deque 
import re
import urllib.parse
from bs4 import BeautifulSoup, Comment # <--- Import Comment
import requests
import time
import html 
import io 
import csv 
import xml.etree.ElementTree as ET # <--- Import XML for Sitemap

# --- IMPORTS FOR NEW FEATURES ---
try:
    from pypdf import PdfReader 
except ImportError:
    print("[!] Failed to import 'pypdf'. Please install: pip3 install pypdf")
    exit()
try:
    import tldextract 
except ImportError:
    print("[!] Failed to import 'tldextract'. Please install: pip3 install tldextract")
    exit()
try:
    from docx import Document # <--- Import for Word
except ImportError:
    print("[!] Failed to import 'python-docx'. Please install: pip3 install python-docx")
    exit()
try:
    from openpyxl import load_workbook # <--- Import for Excel
except ImportError:
    print("[!] Failed to import 'openpyxl'. Please install: pip3 install openpyxl")
    exit()
# --- END IMPORTS ---


HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
}

# Better regex
EMAIL_REGEX = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")

# --- NEW FEATURE: Ignored Extensions Filter ---
IGNORED_EXTENSIONS = [
    '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.svg', '.webp', # Images
    '.css', '.js', # Web Assets
    '.mp4', '.mov', '.avi', '.mp3', '.wav', # Media
    '.zip', '.rar', '.gz', '.tar', # Archives
    'tel:', 'skype:', 'whatsapp:' # Other protocols
]
# --- END NEW FEATURE ---


# --- NEW FEATURE: Sitemap Scanner Function ---
def find_sitemap_urls(base_url):
    """Finds and parses sitemap.xml."""
    sitemap_url = urllib.parse.urljoin(base_url, "/sitemap.xml")
    found_urls = set()
    try:
        response = requests.get(sitemap_url, timeout=10, headers=HEADERS)
        if response.status_code == 200:
            print(f"[i] Found sitemap: {sitemap_url}")
            # Remove XML namespace to make parsing easier
            xml_content = re.sub(r' xmlns="[^"]+"', '', response.text, count=1)
            root = ET.fromstring(xml_content)
            for loc in root.findall('.//loc'):
                if loc.text:
                    found_urls.add(loc.text.strip())
            print(f"    [+] Added {len(found_urls)} URLs from sitemap.")
    except Exception as e:
        print(f"    [!] Could not process sitemap: {e}")
    return found_urls
# --- END NEW FEATURE ---


# --- MAIN SCRIPT ---

user_input = str(input('[+] Enter URL: ')).strip()

if not user_input.startswith('http://') and not user_input.startswith('https://'):
    user_url = 'https://' + user_input
    print(f'[i] Assuming HTTPS. Using: {user_url}')
else:
    user_url = user_input

while True:
    try:
        limit_input = input('[+] How many pages to search (limit)?: ')
        max_urls = int(limit_input) 
        if max_urls > 0: break
        print('[!] Please enter a positive number.')
    except ValueError:
        print('[!] That is not a valid number. Please try again.')

# Determine the root domain
try:
    root_domain_to_stay_on = tldextract.extract(user_url).registered_domain
    print(f"[i] Will only scan links on the root domain: '{root_domain_to_stay_on}' (including subdomains)")
except Exception as e:
    print(f"[!] Failed to extract root domain from {user_url}. Error: {e}")
    exit()

# --- CHANGE: Start queue ---
urls = deque([user_url])
scraped_urls = set()
count = 0
start_time = time.time()
emails = {} # Dictionary for [email]: source_url

# --- NEW FEATURE: Populate queue from sitemap ---
base_sitemap_url = urllib.parse.urlsplit(user_url)._replace(path='').geturl()
sitemap_urls = find_sitemap_urls(base_sitemap_url)
for s_url in sitemap_urls:
    if s_url not in urls and s_url not in scraped_urls:
        urls.append(s_url)
# --- END NEW FEATURE ---

try:
    # Main loop (one-by-one)
    while urls:
        if count >= max_urls: 
            print(f'[!] Reached user-defined limit of {max_urls} pages.')
            break
        
        url = urls.popleft()
        if url in scraped_urls:
            continue
            
        # --- NEW FEATURE: Smart Link Filter ---
        if any(url.lower().endswith(ext) for ext in IGNORED_EXTENSIONS):
            continue
        # --- END NEW FEATURE ---

        scraped_urls.add(url) # Mark as "scraped"
        count += 1
        
        print(f'[{count}/{max_urls}] Processing: {url}')
        time.sleep(0.1) # "Polite" delay
        
        try:
            response = requests.get(url, timeout=15, headers=HEADERS)
            response.raise_for_status() 
            final_url = response.url 
            content_type = response.headers.get('content-type', '').lower()
            new_emails_set = set()
            
            # --- NEW FEATURE: Scan .txt pages ---
            if 'text/plain' in content_type:
                print(f'    [i] Scanning text file: {url}')
                new_emails_set.update(EMAIL_REGEX.findall(response.text))
            
            # --- NEW FEATURE: Scan .docx (Word) ---
            elif 'application/vnd.openxmlformats-officedocument.wordprocessingml.document' in content_type or url.endswith('.docx'):
                print(f'    [i] Scanning Word file: {url}')
                try:
                    with io.BytesIO(response.content) as doc_file:
                        doc = Document(doc_file)
                        doc_text = ""
                        for para in doc.paragraphs:
                            doc_text += para.text + "\n"
                        new_emails_set.update(EMAIL_REGEX.findall(doc_text))
                except Exception: pass
            
            # --- NEW FEATURE: Scan .xlsx (Excel) ---
            elif 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type or url.endswith('.xlsx'):
                print(f'    [i] Scanning Excel file: {url}')
                try:
                    with io.BytesIO(response.content) as xls_file:
                        wb = load_workbook(xls_file, read_only=True)
                        xls_text = ""
                        for sheet in wb.worksheets:
                            for row in sheet.iter_rows():
                                for cell in row:
                                    if cell.value:
                                        xls_text += str(cell.value) + " "
                        new_emails_set.update(EMAIL_REGEX.findall(xls_text))
                except Exception: pass

            # PDF scanning (still here)
            elif 'application/pdf' in content_type:
                print(f'    [i] Scanning PDF file: {url}')
                try:
                    with io.BytesIO(response.content) as pdf_file:
                        reader = PdfReader(pdf_file)
                        pdf_text = ""
                        for page in reader.pages:
                            page_text = page.extract_text()
                            if page_text: pdf_text += page_text
                        new_emails_set.update(EMAIL_REGEX.findall(pdf_text))
                except Exception: pass

            # Process HTML Pages (main logic)
            elif 'text/html' in content_type:
                soup = BeautifulSoup(response.content, 'lxml') # Use the faster 'lxml' parser

                # --- NEW FEATURE: De-obfuscation & Joined Text ---
                # Using .get_text() will join text from separate tags
                html_text = soup.get_text(separator=' ') 
                html_text = html.unescape(html_text)
                html_text = html_text.replace('[at]', '@').replace('(at)', '@')
                html_text = html_text.replace('[dot]', '.').replace('(dot)', '.')
                new_emails_set.update(EMAIL_REGEX.findall(html_text))
                # --- END NEW FEATURE ---

                # --- NEW FEATURE: Scan HTML Comments ---
                comments = soup.find_all(string=lambda text: isinstance(text, Comment))
                for comment in comments:
                    new_emails_set.update(EMAIL_REGEX.findall(comment))
                # --- END NEW FEATURE ---
                
                # Find links
                for a in soup.find_all('a', href=True):
                    href = a['href'].strip()
                    
                    if href.startswith('mailto:'):
                        email = href.replace('mailto:', '').split('?')[0]
                        if email: 
                            new_emails_set.add(email)
                    
                    # Logic for regular links
                    elif not any(href.lower().startswith(ext) for ext in IGNORED_EXTENSIONS):
                        link = urllib.parse.urljoin(final_url, href)
                        
                        # Filter links using SUBDOMAIN
                        try:
                            link_root_domain = tldextract.extract(link).registered_domain
                            if link_root_domain == root_domain_to_stay_on:
                                if link not in scraped_urls and link not in urls:
                                    urls.append(link) # Add to queue
                        except Exception:
                            pass # Ignore broken links

            # --- LOGIC FOR SAVING EMAILS ---
            if new_emails_set:
                new_emails_added_count = 0
                for email in new_emails_set:
                    if email not in emails:
                        emails[email] = url
                        new_emails_added_count += 1
                
                if new_emails_added_count > 0:
                    print(f'    [+] Found {new_emails_added_count} new emails on this page.')

        except requests.exceptions.ReadTimeout:
            print(f'    [!] Server was too slow to respond (timeout): {url}')
            continue
        except (requests.exceptions.RequestException, requests.exceptions.ConnectionError) as e: 
            print(f'    [!] Failed to fetch URL: {url} | Error: {e}')
            continue
            
except KeyboardInterrupt: 
    print('\n[-] Process interrupted by user!')

# --- Finished ---
end_time = time.time()
print(f'\n[+] Process Finished in {end_time - start_time:.2f} seconds.')

# --- OUTPUT FORMAT (still the same) ---
print(f'\nList of Mails ({len(emails)} found):\n==================================')
for mail, source in sorted(emails.items()): 
    print(f'      {mail} (found at: {source})')
print('\n')

# --- SAVING RESULTS TO CSV (still the same) ---
if emails:
    output_filename = 'results.csv'
    print(f'[i] Saving results to {output_filename}...')
    try:
        with open(output_filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Email', 'Source_URL'])
            for mail, source in sorted(emails.items()):
                writer.writerow([mail, source])
        print(f'[+] Successfully saved to {output_filename}')
    except Exception as e:
        print(f'[!] Failed to save file: {e}')
print('\n')
